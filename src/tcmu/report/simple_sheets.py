'''
Module for working with very basic spread sheets.
The main motivation for the creation of this module is
that we cannot programmatically put molecular xyz-coordinates
or any multiline value inside of CSV files. This module
provides a very basic way of including any kind of value inside
a .xlsx file.
'''
from typing import List
import os


def append(file: str, values: List[str]):
    import openpyxl as xl
    '''
    Append a row to a spreadsheets file.

    Args:
        file: the path to the file to append to.
        values: a list of values that will be written into the file.
    '''
    if not os.path.exists(file):
        wb = xl.Workbook()
        ws = wb.active
        row = ws.max_row
    else:
        wb = xl.load_workbook(file)
        ws = wb.active
        row = ws.max_row + 1

    for i, val in enumerate(values):
        ws.cell(row=row, column=i+1, value=val)

    wb.save(file)
